"""Plantilla Excel y exportaciones de perfiles de costeo."""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from services.perfiles_costeo import filas_exportables_combos, filas_exportables_perfiles

COLUMNAS = ["SKU", "TIPO", "UNIDAD", "OBSERVACION"]


def _encabezado_pdf(pdf, titulo, organizacion_nombre, columnas, *, pagina):
    ancho, alto = landscape(A4)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(36, alto - 38, titulo)
    pdf.setFont("Helvetica", 9)
    pdf.drawString(36, alto - 55, organizacion_nombre)
    pdf.drawRightString(ancho - 36, alto - 55, datetime.now().strftime("Emitido %d/%m/%Y %H:%M"))
    pdf.setFillColor(colors.HexColor("#176B89"))
    pdf.rect(36, alto - 88, ancho - 72, 22, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 8)
    for x, texto in columnas:
        pdf.drawString(x, alto - 80, texto)
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(ancho - 36, 22, f"Página {pagina}")
    return alto - 104


def _libro(filas):
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Productos"
    hoja.append(COLUMNAS)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="176B89")
    for fila in filas:
        hoja.append([fila.get("sku", ""), fila.get("tipo", ""), fila.get("unidad", ""), fila.get("observacion", "")])
    for columna, ancho in zip("ABCD", (20, 18, 28, 50)):
        hoja.column_dimensions[columna].width = ancho
    hoja.freeze_panes = "A2"
    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida


def plantilla_excel_productos():
    return _libro([{"sku": "PP6040H", "tipo": "produccion", "unidad": "Fierro 100% Argento", "observacion": "Ejemplo: eliminar esta fila"}])


def exportar_excel_perfiles(perfiles):
    return _libro(filas_exportables_perfiles(perfiles))


def exportar_pdf_perfiles(perfiles, organizacion_nombre):
    salida = BytesIO()
    pdf = canvas.Canvas(salida, pagesize=landscape(A4))
    pagina = 1
    columnas = ((36, "SKU"), (150, "PRODUCTO"), (470, "UNIDAD"), (650, "TIPO"))
    y = _encabezado_pdf(pdf, "Productos y clasificación de costos", organizacion_nombre, columnas, pagina=pagina)
    filas = filas_exportables_perfiles(perfiles)
    if not filas:
        pdf.setFont("Helvetica-Oblique", 10)
        pdf.drawString(36, y, "No hay productos clasificados.")
    for fila in filas:
        if y < 42:
            pdf.showPage()
            pagina += 1
            y = _encabezado_pdf(pdf, "Productos y clasificación de costos", organizacion_nombre, columnas, pagina=pagina)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawString(36, y, str(fila["sku"])[:24])
        pdf.setFont("Helvetica", 9)
        pdf.drawString(150, y, str(fila["producto"])[:60])
        pdf.drawString(470, y, str(fila["unidad"])[:32])
        pdf.drawString(650, y, str(fila["tipo"])[:18])
        y -= 16
    pdf.save()
    salida.seek(0)
    return salida


def plantilla_excel_combos():
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Componentes combos"
    encabezados = ["SKU_COMBO", "SKU_COMPONENTE", "CANTIDAD", "UNIDAD", "OBSERVACION"]
    hoja.append(encabezados)
    hoja.append(["COMBO-1", "PP6040H", 1, "Fierro 100% Argento", "Ejemplo: eliminar esta fila"])
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="176B89")
    for columna, ancho in zip("ABCDE", (20, 24, 14, 28, 50)):
        hoja.column_dimensions[columna].width = ancho
    salida = BytesIO(); libro.save(salida); salida.seek(0)
    return salida


def exportar_excel_combos(perfiles_combo):
    libro = Workbook(); hoja = libro.active; hoja.title = "Componentes combos"
    hoja.append(["SKU_COMBO", "SKU_COMPONENTE", "CANTIDAD", "TIPO_COMPONENTE", "OBSERVACION"])
    for fila in filas_exportables_combos(perfiles_combo):
        hoja.append([fila["sku_combo"], fila["sku_componente"], fila["cantidad"], fila["tipo_componente"], fila["observacion"]])
    salida = BytesIO(); libro.save(salida); salida.seek(0)
    return salida


def exportar_pdf_combos(perfiles_combo, organizacion_nombre):
    salida = BytesIO(); pdf = canvas.Canvas(salida, pagesize=landscape(A4))
    pagina = 1
    columnas = ((36, "COMBO"), (190, "COMPONENTE"), (360, "CANTIDAD"), (470, "TIPO"), (610, "OBSERVACIÓN"))
    y = _encabezado_pdf(pdf, "Componentes de combos", organizacion_nombre, columnas, pagina=pagina)
    filas = filas_exportables_combos(perfiles_combo)
    if not filas:
        pdf.setFont("Helvetica-Oblique", 10); pdf.drawString(36, y, "No hay componentes de combos.")
    for fila in filas:
        if y < 42:
            pdf.showPage(); pagina += 1
            y = _encabezado_pdf(pdf, "Componentes de combos", organizacion_nombre, columnas, pagina=pagina)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(36, y, str(fila["sku_combo"])[:24]); pdf.drawString(190, y, str(fila["sku_componente"])[:28])
        pdf.drawString(360, y, str(fila["cantidad"])); pdf.drawString(470, y, str(fila["tipo_componente"])[:20])
        pdf.drawString(610, y, str(fila["observacion"])[:32]); y -= 16
    pdf.save(); salida.seek(0); return salida
