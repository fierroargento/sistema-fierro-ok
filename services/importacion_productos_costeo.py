"""Lectura, mapeo y validacion masiva de perfiles de costeo."""

import csv
from io import BytesIO, StringIO
import json
import unicodedata

from openpyxl import load_workbook

from services.perfiles_costeo import crear_o_actualizar_perfil


CAMPOS_PRODUCTOS = {
    "sku": {"nombre": "SKU", "obligatorio": True, "alias": {"sku", "codigo", "cod producto"}},
    "tipo": {"nombre": "Tipo de producto", "obligatorio": True, "alias": {"tipo", "tipo producto", "clasificacion"}},
    "unidad": {"nombre": "Unidad de negocio", "obligatorio": False, "alias": {"unidad", "unidad negocio", "marca"}},
    "observacion": {"nombre": "Observacion", "obligatorio": False, "alias": {"observacion", "notas"}},
}
TIPOS_EQUIVALENTES = {
    "simple": "simple", "comprado": "simple",
    "produccion": "produccion", "de produccion": "produccion", "fabricado": "produccion",
    "combo": "combo", "compuesto": "combo",
}


def normalizar(texto):
    return " ".join(
        "".join(c for c in unicodedata.normalize("NFD", str(texto or ""))
                if unicodedata.category(c) != "Mn").lower().strip().split()
    )


def sugerir_mapeo(encabezados):
    resultado = {}
    usados = set()
    for indice, encabezado in enumerate(encabezados):
        limpio = normalizar(encabezado)
        destino = ""
        for campo, definicion in CAMPOS_PRODUCTOS.items():
            if campo not in usados and limpio in definicion["alias"]:
                destino = campo
                usados.add(campo)
                break
        resultado[str(indice)] = destino
    return resultado


def leer_archivo(archivo, nombre_hoja=None, max_filas=5000):
    nombre = str(getattr(archivo, "filename", "") or "").lower()
    contenido = archivo.read()
    if len(contenido) > 10 * 1024 * 1024:
        raise ValueError("El archivo supera el maximo de 10 MB.")
    if nombre.endswith(".csv"):
        texto = contenido.decode("utf-8-sig")
        datos = list(csv.reader(StringIO(texto)))
        hoja = "CSV"
        hojas = [hoja]
    elif nombre.endswith((".xlsx", ".xlsm")):
        libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        hojas = libro.sheetnames
        hoja = nombre_hoja if nombre_hoja in hojas else hojas[0]
        datos = [list(fila) for fila in libro[hoja].iter_rows(values_only=True)]
    else:
        raise ValueError("El archivo debe ser XLSX, XLSM o CSV.")
    if not datos:
        raise ValueError("El archivo no contiene filas.")
    encabezados = [str(valor or "").strip() for valor in datos[0]]
    if not any(encabezados):
        raise ValueError("La primera fila no contiene encabezados.")
    filas = []
    for numero, valores in enumerate(datos[1:max_filas + 1], 2):
        fila = ["" if valor is None else str(valor).strip() for valor in valores]
        if any(fila):
            filas.append({"numero": numero, "valores": fila})
    return {"hoja": hoja, "hojas": hojas, "encabezados": encabezados, "filas": filas}


def extraer_fila(fila, mapeo):
    salida = {}
    for indice, campo in mapeo.items():
        if campo:
            posicion = int(indice)
            salida[campo] = fila["valores"][posicion] if posicion < len(fila["valores"]) else ""
    return salida


def validar_mapeo(mapeo):
    destinos = [campo for campo in mapeo.values() if campo]
    if len(destinos) != len(set(destinos)):
        raise ValueError("Un campo del sistema no puede recibir dos columnas.")
    faltantes = [
        d["nombre"] for campo, d in CAMPOS_PRODUCTOS.items()
        if d["obligatorio"] and campo not in destinos
    ]
    if faltantes:
        raise ValueError("Faltan campos obligatorios: " + ", ".join(faltantes) + ".")


def previsualizar(filas, mapeo, *, organizacion_id, unidad_negocio_id=None, modelos):
    validar_mapeo(mapeo)
    perfiles = modelos["PerfilCosteoProducto"].query.filter_by(
        organizacion_id=organizacion_id
    ).all()
    existentes = {
        ((p.producto.sku or "").upper(), p.unidad_negocio_id): p
        for p in perfiles
    }
    resultado = []
    for fila in filas:
        datos = extraer_fila(fila, mapeo)
        sku = str(datos.get("sku") or "").strip().upper()
        tipo = TIPOS_EQUIVALENTES.get(normalizar(datos.get("tipo")))
        errores = []
        if not sku:
            errores.append("Falta SKU")
        if tipo is None:
            errores.append("Tipo invalido")
        inclusion = None
        if not errores:
            Catalogo = modelos["Catalogo"]
            consulta = modelos["CatalogoProducto"].query.join(Catalogo).join(
                modelos["Producto"]
            ).filter(
                Catalogo.organizacion_id == organizacion_id,
                modelos["Producto"].sku.ilike(sku),
            )
            if unidad_negocio_id is not None:
                consulta = consulta.filter(
                    Catalogo.unidad_negocio_id == unidad_negocio_id
                )
            unidad = normalizar(datos.get("unidad"))
            opciones = consulta.all()
            if unidad:
                opciones = [i for i in opciones if i.catalogo.unidad_negocio and normalizar(i.catalogo.unidad_negocio.nombre) == unidad]
            if len(opciones) != 1:
                errores.append("No se encontro una unica inclusion para SKU y unidad")
            else:
                inclusion = opciones[0]
        actual = existentes.get((
            sku,
            inclusion.catalogo.unidad_negocio_id if inclusion else None,
        ))
        accion = "rechazado" if errores else "actualizar" if actual else "crear"
        if actual and actual.tipo == tipo:
            accion = "sin_cambios"
        resultado.append({
            "numero": fila["numero"], "sku": sku, "tipo": tipo or "",
            "unidad": datos.get("unidad", ""), "observacion": datos.get("observacion", ""),
            "accion": accion, "errores": errores,
            "inclusion_id": inclusion.id if inclusion else None,
        })
    return resultado


def aplicar_vista_previa(vista, *, organizacion_id, modelos, db_session):
    conteos = {"creados": 0, "actualizados": 0, "sin_cambios": 0, "rechazados": 0}
    for fila in vista:
        if fila["accion"] == "rechazado":
            conteos["rechazados"] += 1
            continue
        if fila["accion"] == "sin_cambios":
            conteos["sin_cambios"] += 1
            continue
        inclusion = db_session.get(modelos["CatalogoProducto"], fila["inclusion_id"])
        crear_o_actualizar_perfil(
            organizacion_id=organizacion_id,
            unidad_negocio_id=inclusion.catalogo.unidad_negocio_id,
            producto_id=inclusion.producto_id,
            tipo=fila["tipo"], observacion=fila["observacion"],
            PerfilCosteoProducto=modelos["PerfilCosteoProducto"],
            UnidadNegocio=modelos["UnidadNegocio"], Producto=modelos["Producto"],
            db_session=db_session, commit=False,
        )
        conteos["creados" if fila["accion"] == "crear" else "actualizados"] += 1
    try:
        db_session.commit()
    except Exception:
        db_session.rollback()
        raise
    return conteos


def serializar(valor):
    return json.dumps(valor, ensure_ascii=False, default=str)


def deserializar(valor, predeterminado):
    return json.loads(valor) if valor else predeterminado
