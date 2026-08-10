"""Contratos separados para carga masiva de fuentes y fichas productivas."""

from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from services.importacion_productos_costeo import normalizar

try:
    from openpyxl.styles import Font, PatternFill
except ImportError:  # El runtime minimo de pruebas no expone estilos.
    Font = PatternFill = None


DEFINICIONES = {
    "insumos": {
        "titulo": "Insumos y precios",
        "campos": {
            "codigo": ("Código", True), "nombre": ("Nombre", True),
            "tipo": ("Tipo", True), "unidad_medida": ("Unidad de medida", True),
            "precio_unitario": ("Precio unitario", True), "proveedor": ("Proveedor", False),
        },
        "ejemplo": ["hierro-6mm", "Hierro redondo 6 mm", "materia_prima", "kg", "1500", "Proveedor"],
    },
    "empleados": {
        "titulo": "Empleados y costos laborales",
        "campos": {
            "codigo": ("Código o legajo", True), "nombre": ("Nombre", True),
            "sector": ("Sector", True), "puesto": ("Puesto", False),
            "ubicacion_trabajo": ("Ubicación de trabajo", True),
            "tipo_funcion": ("Tipo de función", True),
            "porcentaje_productivo": ("Participación productiva %", True),
            "sueldo_base": ("Sueldo base", True),
            "porcentaje_cargas": ("Excepción cargas %", False),
            "adicionales": ("Adicionales", False), "otros_costos": ("Otros costos", False),
            "horas_mensuales": ("Horas mensuales", True), "horas_productivas": ("Horas productivas", True),
        },
        "ejemplo": ["EMP-1", "Operario", "Herrería", "Soldador", "Taller", "directa", "100", "1000000", "", "0", "0", "176", "160"],
    },
    "recursos": {
        "titulo": "Recursos y equipos de mano de obra",
        "campos": {
            "codigo_recurso": ("Código del recurso", True),
            "nombre_recurso": ("Nombre del recurso", True),
            "sector": ("Sector", True),
            "porcentaje_indirecto": ("Tiempo indirecto %", True),
            "codigo_empleado": ("Código del empleado", True),
            "porcentaje_dedicacion": ("Dedicación %", True),
        },
        "ejemplo": ["HERRERIA", "Equipo Herrería", "Herrería", "10", "EMP-1", "100"],
    },
    "costos-fijos": {
        "titulo": "Costos fijos",
        "campos": {
            "codigo": ("Código", True), "nombre": ("Nombre", True),
            "categoria": ("Categoría", True), "integra_produccion": ("Integra producción", True),
            "criterio": ("Criterio de distribución", True), "importe_mensual": ("Importe mensual", True),
            "comprobante": ("Comprobante", False),
        },
        "ejemplo": ["alquiler", "Alquiler del galpón", "Infraestructura", "si", "unidades_producidas", "500000", "Factura"],
    },
    "fichas": {
        "titulo": "Componentes de fichas técnicas",
        "campos": {
            "sku": ("SKU producto", True), "tipo_linea": ("Tipo de línea", True),
            "codigo_recurso": ("Código del recurso", True), "cantidad": ("Cantidad", False),
            "merma": ("Merma %", False), "operacion": ("Operación", False),
            "minutos": ("Minutos", False), "porcentaje": ("Asignación %", False),
            "unidades_mensuales": ("Unidades mensuales", False),
        },
        "ejemplo": ["PP6040H", "insumo", "hierro-6mm", "2.5", "5", "", "", "", ""],
    },
}


ETIQUETAS_VALORES = {
    "materia_prima": "Materia prima",
    "consumible": "Consumible",
    "servicio_productivo": "Servicio productivo",
    "embalaje_productivo": "Embalaje productivo",
    "insumo": "Insumo",
    "operacion": "Operación",
    "costo_fijo": "Costo fijo",
    "fijo": "Costo fijo",
    "unidades_producidas": "Unidades producidas",
    "horas_productivas": "Horas productivas",
    "si": "Sí",
    "sí": "Sí",
    "no": "No",
}

CAMPOS_MONEDA = {
    "precio_unitario", "sueldo_base", "adicionales",
    "otros_costos", "importe_mensual",
}

CAMPOS_PORCENTAJE = {
    "merma", "porcentaje", "porcentaje_indirecto", "porcentaje_dedicacion",
    "porcentaje_cargas", "porcentaje_productivo",
}

TIPOS_INSUMO_VALIDOS = {
    "materia_prima", "consumible", "servicio_productivo", "embalaje_productivo",
}

CRITERIOS_DISTRIBUCION_VALIDOS = {
    "horas_productivas", "horas_maquina", "unidades_producidas",
    "porcentaje", "importe_directo", "sin_distribuir",
}


def definicion(tipo):
    if tipo not in DEFINICIONES:
        raise ValueError("El tipo de importación no es válido.")
    base = DEFINICIONES[tipo]
    return {
        **base,
        "campos_ui": {
            clave: {"nombre": datos[0], "obligatorio": datos[1]}
            for clave, datos in base["campos"].items()
        },
    }


def sugerir_mapeo_fuente(tipo, encabezados):
    campos = definicion(tipo)["campos"]
    resultado, usados = {}, set()
    for indice, encabezado in enumerate(encabezados):
        limpio = normalizar(str(encabezado or "").replace("_", " ").replace("-", " "))
        destino = ""
        for clave, (nombre, _obligatorio) in campos.items():
            opciones = {normalizar(nombre), normalizar(clave.replace("_", " "))}
            if clave not in usados and limpio in opciones:
                destino, usados = clave, usados | {clave}
                break
        resultado[str(indice)] = destino
    return resultado


def _formatear_decimal(valor, decimales=2):
    try:
        numero = Decimal(str(valor).replace(",", ".").strip())
    except (InvalidOperation, ValueError, AttributeError):
        return str(valor or "")
    texto = f"{numero:,.{decimales}f}"
    return texto.replace(",", "_").replace(".", ",").replace("_", ".")


def _valor_presentado(campo, valor):
    if valor is None or str(valor).strip() == "":
        return "—"
    limpio = normalizar(valor)
    if campo in CAMPOS_MONEDA:
        return f"$ {_formatear_decimal(valor)}"
    if campo in CAMPOS_PORCENTAJE:
        return f"{_formatear_decimal(valor)} %"
    return ETIQUETAS_VALORES.get(limpio, str(valor))


def presentar_vista_fuentes(tipo, vista):
    """Prepara etiquetas legibles sin alterar los datos confirmables del lote."""
    campos = definicion(tipo)["campos"]
    columnas = [{"clave": clave, "nombre": nombre} for clave, (nombre, _obligatorio) in campos.items()]
    acciones = {
        "crear": ("Crear", "create"),
        "actualizar": ("Actualizar", "update"),
        "rechazado": ("Rechazado", "rejected"),
    }
    filas = []
    for fila in vista:
        etiqueta, clase = acciones.get(fila.get("accion"), (str(fila.get("accion") or ""), "draft"))
        datos = fila.get("datos") or {}
        filas.append({
            "numero": fila.get("numero"),
            "valores": [_valor_presentado(columna["clave"], datos.get(columna["clave"])) for columna in columnas],
            "accion": etiqueta,
            "clase_accion": clase,
            "detalle": ", ".join(fila.get("errores") or []),
        })
    return columnas, filas


def aplicar_modo_vista_fuentes(vista, modo):
    """Aplica el alcance elegido sin modificar las filas originales."""
    resultado = []
    for fila_original in vista:
        fila = {
            **fila_original,
            "errores": list(fila_original.get("errores") or []),
        }
        if modo == "solo_crear" and fila.get("accion") == "actualizar":
            fila["accion"] = "rechazado"
            fila["errores"] = ["El registro ya existe"]
        elif modo == "solo_actualizar" and fila.get("accion") == "crear":
            fila["accion"] = "rechazado"
            fila["errores"] = ["El registro todavía no existe"]
        resultado.append(fila)
    return resultado


def resumir_vista_fuentes(vista):
    resumen = {"crear": 0, "actualizar": 0, "rechazado": 0, "aplicables": 0}
    for fila in vista:
        accion = str(fila.get("accion") or "")
        if accion in resumen:
            resumen[accion] += 1
        if accion in {"crear", "actualizar"}:
            resumen["aplicables"] += 1
    return resumen


def _extraer(fila, mapeo):
    return {
        campo: (fila["valores"][int(indice)] if int(indice) < len(fila["valores"]) else "")
        for indice, campo in mapeo.items() if campo
    }


def _numero(valor, nombre, obligatorio=True):
    if not str(valor or "").strip() and not obligatorio:
        return "0"
    try:
        numero = Decimal(str(valor).replace(",", ".").strip())
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{nombre} no es válido") from error
    if not numero.is_finite() or numero < 0:
        raise ValueError(f"{nombre} no es válido")
    return str(numero)


def _clave_valor(valor):
    return normalizar(valor).replace(" ", "_").replace("-", "_")


def _texto_requerido(datos, campo, nombre):
    texto = str(datos.get(campo) or "").strip()
    if not texto:
        raise ValueError(f"{nombre} es obligatorio")
    datos[campo] = texto
    return texto


def _numero_positivo(valor, nombre, maximo=None):
    numero = Decimal(_numero(valor, nombre))
    if numero <= 0 or (maximo is not None and numero > Decimal(str(maximo))):
        raise ValueError(f"{nombre} está fuera de rango")
    return str(numero)


def previsualizar_fuentes(tipo, filas, mapeo, *, organizacion_id, unidad_negocio_id, modelos):
    config = definicion(tipo)
    destinos = [campo for campo in mapeo.values() if campo]
    faltantes = [nombre for clave, (nombre, obligatorio) in config["campos"].items() if obligatorio and clave not in destinos]
    if faltantes:
        raise ValueError("Faltan campos obligatorios: " + ", ".join(faltantes) + ".")
    if len(destinos) != len(set(destinos)):
        raise ValueError("Un campo del sistema no puede recibir dos columnas.")
    resultado, identidades = [], set()
    for fila in filas:
        datos, errores, existente, ids = _extraer(fila, mapeo), [], None, {}
        vinculo = None
        try:
            for clave, (nombre, obligatorio) in config["campos"].items():
                if obligatorio:
                    _texto_requerido(datos, clave, nombre)
            if tipo == "insumos":
                datos["tipo"] = _clave_valor(datos.get("tipo"))
                if datos["tipo"] not in TIPOS_INSUMO_VALIDOS:
                    raise ValueError("Tipo de insumo inválido")
                existente = modelos["InsumoProductivo"].query.filter_by(organizacion_id=organizacion_id, codigo=str(datos.get("codigo") or "").strip().lower()).first()
                if existente and existente.unidad_negocio_id not in {None, unidad_negocio_id}:
                    raise ValueError("El código ya pertenece a otra unidad")
                if existente: ids = {"insumo_id": existente.id}
                datos["precio_unitario"] = _numero(datos.get("precio_unitario"), "Precio unitario")
            elif tipo == "empleados":
                existente = modelos["EmpleadoProductivo"].query.filter_by(organizacion_id=organizacion_id, codigo=str(datos.get("codigo") or "").strip().lower()).first()
                if existente and existente.unidad_negocio_id not in {None, unidad_negocio_id}:
                    raise ValueError("El código ya pertenece a otra unidad")
                if existente: ids = {"empleado_id": existente.id}
                datos["tipo_funcion"] = _clave_valor(datos.get("tipo_funcion"))
                if datos["tipo_funcion"] not in {
                    "directa", "indirecta_productiva",
                    "comercial_administrativa", "mixta",
                }:
                    raise ValueError("Tipo de función laboral inválido")
                datos["porcentaje_productivo"] = _numero(
                    datos.get("porcentaje_productivo"), "Participación productiva",
                )
                if Decimal(datos["porcentaje_productivo"]) > Decimal("100"):
                    raise ValueError("Participación productiva está fuera de rango")
                for campo in ("sueldo_base", "adicionales", "otros_costos", "horas_mensuales", "horas_productivas"):
                    datos[campo] = _numero(datos.get(campo), campo, campo in {"sueldo_base", "horas_mensuales", "horas_productivas"})
                excepcion = str(datos.get("porcentaje_cargas") or "").strip()
                datos["porcentaje_cargas"] = (
                    _numero(excepcion, "Excepción de cargas") if excepcion else ""
                )
                if excepcion and Decimal(datos["porcentaje_cargas"]) > Decimal("100"):
                    raise ValueError("Excepción de cargas está fuera de rango")
                datos["horas_mensuales"] = _numero_positivo(datos["horas_mensuales"], "Horas mensuales")
                datos["horas_productivas"] = _numero_positivo(datos["horas_productivas"], "Horas productivas")
                if Decimal(datos["horas_productivas"]) > Decimal(datos["horas_mensuales"]):
                    raise ValueError("Las horas productivas no pueden superar las horas mensuales")
            elif tipo == "recursos":
                codigo_recurso = str(datos.get("codigo_recurso") or "").strip().lower()
                codigo_empleado = str(datos.get("codigo_empleado") or "").strip().lower()
                existente = modelos["EmpleadoProductivo"].query.filter_by(
                    organizacion_id=organizacion_id,
                    codigo=codigo_recurso,
                ).first()
                if existente and existente.tipo_registro != "recurso":
                    raise ValueError("El código del recurso corresponde a un empleado")
                if existente and existente.unidad_negocio_id != unidad_negocio_id:
                    raise ValueError("El recurso pertenece a otra unidad")
                empleado = modelos["EmpleadoProductivo"].query.filter_by(
                    organizacion_id=organizacion_id,
                    codigo=codigo_empleado,
                    tipo_registro="empleado",
                ).first()
                if empleado is None or empleado.unidad_negocio_id != unidad_negocio_id:
                    raise ValueError("No se encontró el empleado en la unidad activa")
                datos["porcentaje_indirecto"] = _numero(
                    datos.get("porcentaje_indirecto"), "Tiempo indirecto",
                )
                datos["porcentaje_dedicacion"] = _numero_positivo(
                    datos.get("porcentaje_dedicacion"), "Dedicación", 100,
                )
                if Decimal(datos["porcentaje_indirecto"]) > Decimal("100"):
                    raise ValueError("Tiempo indirecto está fuera de rango")
                ids = {"empleado_id": empleado.id}
                if existente:
                    ids["recurso_id"] = existente.id
                    vinculo = modelos["RecursoEmpleadoProductivo"].query.filter_by(
                        recurso_id=existente.id, empleado_id=empleado.id,
                    ).first()
                    if vinculo:
                        ids["vinculo_id"] = vinculo.id
                existente = vinculo
            elif tipo == "costos-fijos":
                integra = _clave_valor(datos.get("integra_produccion"))
                equivalencias = {"si": "si", "1": "si", "true": "si", "no": "no", "0": "no", "false": "no"}
                if integra not in equivalencias:
                    raise ValueError("Integra producción debe indicar sí o no")
                datos["integra_produccion"] = equivalencias[integra]
                datos["criterio"] = _clave_valor(datos.get("criterio"))
                if datos["criterio"] not in CRITERIOS_DISTRIBUCION_VALIDOS:
                    raise ValueError("Criterio de distribución inválido")
                if datos["integra_produccion"] == "no" and datos["criterio"] != "sin_distribuir":
                    raise ValueError("Un costo informativo debe quedar sin distribuir")
                existente = modelos["CostoFijoProductivo"].query.filter_by(organizacion_id=organizacion_id, codigo=str(datos.get("codigo") or "").strip().lower()).first()
                if existente and existente.unidad_negocio_id not in {None, unidad_negocio_id}:
                    raise ValueError("El código ya pertenece a otra unidad")
                if existente: ids = {"costo_fijo_id": existente.id}
                datos["importe_mensual"] = _numero(datos.get("importe_mensual"), "Importe mensual")
            elif tipo == "fichas":
                sku = str(datos.get("sku") or "").strip().upper()
                perfil = modelos["PerfilCosteoProducto"].query.join(modelos["Producto"]).filter(
                    modelos["PerfilCosteoProducto"].organizacion_id == organizacion_id,
                    modelos["PerfilCosteoProducto"].unidad_negocio_id == unidad_negocio_id,
                    modelos["PerfilCosteoProducto"].tipo == "produccion",
                    modelos["Producto"].sku.ilike(sku),
                ).first()
                if perfil is None: raise ValueError("No se encontró el producto de producción")
                linea = normalizar(datos.get("tipo_linea")).replace("_", " ")
                datos["tipo_linea"] = "costo_fijo" if linea in {"costo fijo", "fijo"} else linea
                codigo = str(datos.get("codigo_recurso") or "").strip().lower()
                mapa = {"insumo": ("InsumoProductivo", "insumo_id"), "operacion": ("EmpleadoProductivo", "empleado_id"), "costo fijo": ("CostoFijoProductivo", "costo_fijo_id"), "fijo": ("CostoFijoProductivo", "costo_fijo_id")}
                if linea not in mapa: raise ValueError("Tipo de línea inválido")
                nombre_modelo, campo_id = mapa[linea]
                recurso = modelos[nombre_modelo].query.filter_by(organizacion_id=organizacion_id, codigo=codigo).first()
                if recurso is None or recurso.unidad_negocio_id not in {None, unidad_negocio_id}: raise ValueError("No se encontró el recurso en la unidad activa")
                ids = {"perfil_costeo_id": perfil.id, campo_id: recurso.id}
                if linea == "insumo":
                    datos["cantidad"] = _numero_positivo(datos.get("cantidad"), "Cantidad")
                    datos["merma"] = _numero(datos.get("merma"), "Merma", False)
                    if Decimal(datos["merma"]) > Decimal("100"):
                        raise ValueError("Merma está fuera de rango")
                    existente = next((x for x in perfil.insumos_costeo if x.insumo_id == recurso.id), None)
                elif linea in {"costo fijo", "fijo"}:
                    if not recurso.integra_costo_produccion:
                        raise ValueError("El costo fijo no está marcado como productivo")
                    datos["porcentaje"] = _numero_positivo(datos.get("porcentaje"), "Asignación %", 100)
                    datos["unidades_mensuales"] = _numero_positivo(datos.get("unidades_mensuales"), "Unidades mensuales")
                    existente = next((x for x in perfil.costos_fijos_costeo if x.costo_fijo_id == recurso.id), None)
                else:
                    _texto_requerido(datos, "operacion", "Operación")
                    datos["minutos"] = _numero_positivo(datos.get("minutos"), "Minutos")
                    existente = next((x for x in perfil.operaciones_costeo if x.empleado_id == recurso.id and normalizar(x.nombre) == normalizar(datos["operacion"])), None)
                    if existente: ids["operacion_id"] = existente.id
            identidad = (
                tipo, str(datos.get("codigo") or datos.get("codigo_recurso") or datos.get("sku") or "").strip().lower(),
                str(datos.get("tipo_linea") or ""), str(datos.get("codigo_recurso") or "").strip().lower(),
                normalizar(datos.get("operacion")) if tipo == "fichas" else str(datos.get("codigo_empleado") or "").strip().lower(),
            )
            if identidad in identidades:
                raise ValueError("El archivo contiene una fila duplicada para el mismo registro")
            identidades.add(identidad)
        except ValueError as error:
            errores.append(str(error))
        resultado.append({"numero": fila["numero"], "datos": datos, "ids": ids, "accion": "rechazado" if errores else "actualizar" if existente else "crear", "errores": errores})
    return resultado


def plantilla_excel_fuente(tipo):
    config = definicion(tipo); libro = Workbook(); hoja = libro.active
    hoja.title = config["titulo"][:31]
    hoja.append([nombre.upper() for nombre, _ in config["campos"].values()]); hoja.append(config["ejemplo"])
    if Font is not None and PatternFill is not None:
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="176B89")
    hoja.freeze_panes = "A2"
    for columna in hoja.columns:
        hoja.column_dimensions[columna[0].column_letter].width = max(16, min(32, len(str(columna[0].value)) + 4))
    salida = BytesIO(); libro.save(salida); salida.seek(0); return salida


def aplicar_fuentes(tipo, vista, *, organizacion, unidad_activa, modelos, db_session, usuario):
    from services.fuentes_costo_admin import procesar_accion_fuente_costo
    conteos = {"creados": 0, "actualizados": 0, "sin_cambios": 0, "rechazados": 0}
    for fila in vista:
        if fila["accion"] == "rechazado":
            conteos["rechazados"] += 1; continue
        datos = dict(fila["datos"]); datos.update(fila.get("ids") or {})
        datos["unidad_negocio_id"] = str(unidad_activa.id)
        if tipo == "insumos":
            accion = "crear_insumo" if fila["accion"] == "crear" else "actualizar_precio_insumo"
            datos.update({"insumo_id": datos.get("insumo_id"), "precio_unitario": datos.get("precio_unitario"), "proveedor_referencia": datos.get("proveedor")})
        elif tipo == "empleados":
            accion = "crear_empleado" if fila["accion"] == "crear" else "actualizar_costo_empleado"
            datos["empleado_id"] = datos.get("empleado_id")
        elif tipo == "recursos":
            recurso = modelos["EmpleadoProductivo"].query.filter_by(
                organizacion_id=organizacion.id,
                codigo=str(datos.get("codigo_recurso") or "").strip().lower(),
            ).first()
            if recurso is None:
                procesar_accion_fuente_costo(
                    "crear_recurso_productivo", {
                        "unidad_negocio_id": str(unidad_activa.id),
                        "codigo": datos.get("codigo_recurso"),
                        "nombre": datos.get("nombre_recurso"),
                        "sector": datos.get("sector"),
                        "porcentaje_indirecto": datos.get("porcentaje_indirecto"),
                    }, organizacion=organizacion, unidad_activa=unidad_activa,
                    modelos=modelos, db_session=db_session, usuario=usuario,
                )
                recurso = modelos["EmpleadoProductivo"].query.filter_by(
                    organizacion_id=organizacion.id,
                    codigo=str(datos.get("codigo_recurso") or "").strip().lower(),
                ).first()
            else:
                recurso.nombre = str(datos.get("nombre_recurso") or "").strip()
                recurso.sector = str(datos.get("sector") or "").strip()
                recurso.porcentaje_indirecto = Decimal(datos.get("porcentaje_indirecto") or "0")
                db_session.commit()
            procesar_accion_fuente_costo(
                "vincular_empleado_recurso", {
                    "recurso_id": recurso.id,
                    "empleado_id": datos.get("empleado_id"),
                    "porcentaje_dedicacion": datos.get("porcentaje_dedicacion"),
                }, organizacion=organizacion, unidad_activa=unidad_activa,
                modelos=modelos, db_session=db_session, usuario=usuario,
            )
            conteos["creados" if fila["accion"] == "crear" else "actualizados"] += 1
            continue
        elif tipo == "costos-fijos":
            accion = "crear_costo_fijo" if fila["accion"] == "crear" else "actualizar_importe_costo_fijo"
            datos.update({"costo_fijo_id": datos.get("costo_fijo_id"), "integra_costo_produccion": "1" if normalizar(datos.get("integra_produccion")) in {"si", "sí", "1", "true"} else "0", "criterio_distribucion": datos.get("criterio"), "comprobante_referencia": datos.get("comprobante")})
        else:
            linea = normalizar(datos.get("tipo_linea"))
            accion = "ficha_insumo" if linea == "insumo" else "ficha_operacion" if linea == "operacion" else "ficha_costo_fijo"
            datos.update({"merma": datos.get("merma", 0), "nombre_operacion": datos.get("operacion"), "porcentaje": datos.get("porcentaje")})
        procesar_accion_fuente_costo(
            accion, datos, organizacion=organizacion, unidad_activa=unidad_activa,
            modelos=modelos, db_session=db_session, usuario=usuario,
        )
        conteos["creados" if fila["accion"] == "crear" else "actualizados"] += 1
    return conteos


def exportar_excel_tabla(titulo, encabezados, filas):
    libro = Workbook(); hoja = libro.active; hoja.title = titulo[:31]; hoja.append(encabezados)
    for fila in filas: hoja.append(fila)
    if Font is not None and PatternFill is not None:
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="176B89")
    hoja.freeze_panes = "A2"; salida = BytesIO(); libro.save(salida); salida.seek(0); return salida


def exportar_pdf_tabla(titulo, unidad, encabezados, filas):
    salida = BytesIO(); pdf = canvas.Canvas(salida, pagesize=landscape(A4)); _ancho, alto = landscape(A4); y = alto - 38
    pdf.setFont("Helvetica-Bold", 15); pdf.drawString(32, y, titulo); y -= 18
    pdf.setFont("Helvetica", 9); pdf.drawString(32, y, unidad); y -= 22
    anchos = [32 + indice * (770 / max(1, len(encabezados))) for indice in range(len(encabezados))]
    pdf.setFont("Helvetica-Bold", 7)
    for x, texto in zip(anchos, encabezados): pdf.drawString(x, y, str(texto)[:20])
    y -= 15; pdf.setFont("Helvetica", 7)
    if not filas: pdf.drawString(32, y, "Sin registros.")
    for fila in filas:
        if y < 35: pdf.showPage(); y = alto - 38
        for x, valor in zip(anchos, fila): pdf.drawString(x, y, str(valor or "")[:22])
        y -= 13
    pdf.save(); salida.seek(0); return salida
